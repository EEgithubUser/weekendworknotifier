import datetime
import subprocess
import os
from datetime import date
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side

working = []
play_on = True

while play_on:
	print('\n'*100)
	print(f"You have selected the following people: {working}\n \n")
	try:
		ask = int(input("Which of the following people will be working this weekend? \n 0: Remove Last Worker \n 1. Worker 1 \n 2. Worker 2 \n 3. Worker 3 \n 4. Worker 4 \n 5. Worker 5 \n 6. Worker 6 \n 7. Worker 7 \n 8. Worker 8 \n 9. Worker 9 \n 10. Worker 10 \n 11. Worker 11 \n 12. Worker 12 \n 13. Worker 13 \n 14. Export to Excel \n\n"))
		if ask == 14:
			play_on = False
			print('\n'*100)
			print(f"Working this weekend: {working}")
		elif ask == 1:
			working.append("Worker 1")
		elif ask == 2:
			working.append("Worker 2")
		elif ask == 3:
			working.append("Worker 3")
		elif ask == 4:
			working.append("Worker 4")
		elif ask == 5:
			working.append("Worker 5")
		elif ask == 6:
			working.append("Worker 6")
		elif ask == 7:
			working.append("Worker 7")
		elif ask == 8:
			working.append("Worker 8")
		elif ask == 9:
			working.append("Worker 9")
		elif ask == 10:
			working.append("Worker 10")
		elif ask == 11:
			working.append("Worker 11")
		elif ask == 12:
			working.append("Worker 12")
		elif ask == 13:
			working.append("Worker 13")
		elif ask == 0:
			working.pop()				
		else:
			print("Please enter a number from 0-14")
	except:
			print("Please enter the correct number!")
			continue

# Find date of coming Saturday and Sunday
def get_next_weekday(startdate, weekday):
	d = datetime.strptime(startdate, '%Y-%m-%d')
	t = timedelta((7 + weekday - d.weekday()) % 7)
	
	return (d + t).strftime("%#m/%#d/%y")
	
day = date.today()
date = day.strftime("%Y-%m-%d")

saturday = get_next_weekday(date, 5)
sunday = get_next_weekday(date,6)

# Get Saturday and Sunday date
def get_next_weekday2(startdate, weekday):
	d = datetime.strptime(startdate, '%Y-%m-%d')
	t = timedelta((7 + weekday - d.weekday()) % 7)
	
	return (d + t).strftime("%B %d")

start = get_next_weekday2(date, 5)
end = get_next_weekday2(date,6)

# Excel Sheet Formatting
fontStyle = Font(size = "13")
fontDefault = Font(size = "11")

wb = load_workbook('Weekend Work Automation.xlsx')
ws1 = wb.create_sheet("pop")
wb.active = wb["pop"]
ws = wb.active
ws.title = f"{start} - {end}" 
ws.sheet_view.zoomScale = 150 # Set zoom to 150% for the worksheet we just created

ws['A1'].alignment = Alignment(wrap_text=True)
ws['B1'].alignment = Alignment(wrap_text=True)
ws['C1'].alignment = Alignment(wrap_text=True)
ws.cell(row = 1, column =1, value = 'Weekend Work').font = fontStyle
ws.cell(row = 1, column =2, value = f'Saturday \n({saturday})').font = fontDefault
ws.cell(row = 1, column =3, value = f'Sunday \n({sunday})').font = fontDefault

ws.row_dimensions[1].height = 43.80
ws.column_dimensions['A'].width = 16.90

# Set Names
ws['A2'].value = "Worker 1"
ws['A3'].value = "Worker 2"
ws['A4'].value = "Worker 3"
ws['A5'].value = "Worker 4"
ws['A6'].value = "Worker 5"
ws['A7'].value = "Worker 6"
ws['A8'].value = "Worker 7"
ws['A9'].value = "Worker 8"
ws['A10'].value = "Worker 9"
ws['A11'].value = "Worker 10"
ws['A12'].value = "Worker 11"
ws['A13'].value = "Worker 12"
ws['A14'].value = "Worker 13"

# Set ALL Default Off Days
def set_off_days():
	for r in range(1,14):
		for c in range(1,3):
			ws.cell(row=1+r, column=1+c).value = 'OFF'

set_off_days()

# Set Boarders
def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

set_border(ws, 'A1:C14')

# Take names from working list and mark WORK or OFF
# If "name" is found in list mark WORK else mark OFF
count = len(working)

def set_working():
	for name in working[0:count]:
		if name == "Worker 1":
			ws['B2'].value = "WORK"
		elif name == "Worker 1":
			ws['B3'].value = "WORK"
		elif name == "Worker 1":
			ws['B4'].value = "WORK"
		elif name == "Worker 1":
			ws['B5'].value = "WORK"
		elif name == "Worker 1":
			ws['B6'].value = "WORK"
		elif name == "Worker 1":
			ws['B7'].value = "WORK"
		elif name == "Worker 1":
			ws['B8'].value = "WORK"
		elif name == "Worker 1":
			ws['C9'].value = "WORK"
		elif name == "Worker 1":
			ws['C10'].value = "WORK"	
		elif name == "Worker 1":
			ws['C11'].value = "WORK"
		elif name == "Worker 1":
			ws['C12'].value = "WORK"
		elif name == "Worker 1":
			ws['C13'].value = "WORK"
		elif name == "Worker 1":
			ws['C14'].value = "WORK"
		else:
			pass

set_working()

# Create message to copy
ws['A16'].value = f"Weekend Work - ({saturday} - {sunday})"
ws['A17'].value = f"Hi,\n\nThe following people will be working this weekend ({saturday} - {sunday})\n\n"
wb.save('Weekend Work Automation.xlsx')

# Open Excel Sheet!!
file = 'Weekend Work Automation.xlsx'
os.startfile(file)