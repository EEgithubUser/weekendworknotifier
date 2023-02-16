import datetime
import subprocess
import os
from configparser import ConfigParser
from datetime import date
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side

class RunProgram():

	def config_setup(self):
		file = 'config.ini'
		self.config = ConfigParser()
		self.config.read(file)


	def create_working_list(self):
		self.working_list = []


	def populate_working_list(self):
		PROMPT = '''
Which of the following people will be working this weekend? 
0: Remove Last User
1. Worker 1
2. Worker 2
3. Worker 3
4. Worker 4
5. Worker 5
6. Worker 6
7. Worker 7
8. Worker 8
9. Worker 9
10. Worker 10
11. Worker 11
12. Worker 12
13. Worker 13
q. QUIT
'''

		is_running = True

		while is_running:
			print('\n'*100)
			print(f"You have selected the following people: {self.working_list}\n \n")
			try:
				ask = input(PROMPT)
				if ask == 'q':
					is_running = False
					print('\n'*100)
					print(f"Working this weekend: {self.working_list}")
				elif ask == '1':
					self.working_list.append("Worker 1")
				elif ask == '2':
					self.working_list.append("Worker 2")
				elif ask == '3':
					self.working_list.append("Worker 3")
				elif ask == '4':
					self.working_list.append("Worker 4")
				elif ask == '5':
					self.working_list.append("Worker 5")
				elif ask == '6':
					self.working_list.append("Worker 6")
				elif ask == '7':
					self.working_list.append("Worker 7")
				elif ask == '8':
					self.working_list.append("Worker 8")
				elif ask == '9':
					self.working_list.append("Worker 9")
				elif ask == '10':
					self.working_list.append("Worker 10")
				elif ask == '11':
					self.working_list.append("Worker 11")
				elif ask == '12':
					self.working_list.append("Worker 12")
				elif ask == '13':
					self.working_list.append("Worker 13")
				elif ask == '0':
					self.working_list.pop()				
				else:
					print("Please enter the corresponding character from the prompt.")
			except:
					print("Invalid input.")
					continue

	def get_next_weekday(self, startdate, weekday):
		d = datetime.strptime(startdate, '%Y-%m-%d')
		t = timedelta((7 + weekday - d.weekday()) % 7)
		
		return (d + t)


	def get_current_date(self):
		self.day = date.today()
		self.date = self.day.strftime("%Y-%m-%d")


	def get_next_weekend(self):
		self.saturday = self.get_next_weekday(self.date, 5).strftime("%#m/%#d/%y")
		self.sunday = self.get_next_weekday(self.date,6).strftime("%#m/%#d/%y")


	def format_wkst_title_dates(self):
		self.sat_wkst_title_date = self.get_next_weekday(self.date, 5).strftime("%B %d")
		self.sun_wkst_title_date = self.get_next_weekday(self.date,6).strftime("%B %d")


	def set_border(self, ws, cell_range):
	    thin = Side(border_style="thin", color="000000")
	    for row in ws[cell_range]:
	        for cell in row:
	            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


	def set_off_days(self):
		for r in range(1,14):
			for c in range(1,3):
				self.ws.cell(row=1+r, column=1+c).value = 'OFF'


	def populate_cells(self):
		self.ws['A2'].value = "Worker 1"
		self.ws['A3'].value = "Worker 2"
		self.ws['A4'].value = "Worker 3"
		self.ws['A5'].value = "Worker 4"
		self.ws['A6'].value = "Worker 5"
		self.ws['A7'].value = "Worker 6"
		self.ws['A8'].value = "Worker 7"
		self.ws['A9'].value = "Worker 8"
		self.ws['A10'].value = "Worker 9"
		self.ws['A11'].value = "Worker 10"
		self.ws['A12'].value = "Worker 11"
		self.ws['A13'].value = "Worker 12"
		self.ws['A14'].value = "Worker 13"
		self.ws['A16'].value = f"Weekend Work ({self.saturday} - {self.sunday})"
		self.ws['A17'].value = f"Hi,\n\nThe following people will be working this weekend ({self.saturday} - {self.sunday})\n\n"


	def set_working(self):
		count = len(self.working_list)
		
		for name in self.working_list[0:count]:
			if name == "Worker 1":
				self.ws['B2'].value = "WORK"
			elif name == "Worker 2":
				self.ws['B3'].value = "WORK"
			elif name == "Worker 3":
				self.ws['B4'].value = "WORK"
			elif name == "Worker 4":
				self.ws['B5'].value = "WORK"
			elif name == "Worker 5":
				self.ws['B6'].value = "WORK"
			elif name == "Worker 6":
				self.ws['B7'].value = "WORK"
			elif name == "Worker 7":
				self.ws['B8'].value = "WORK"
			elif name == "Worker 8":
				self.ws['C9'].value = "WORK"
			elif name == "Worker 9":
				self.ws['C10'].value = "WORK"	
			elif name == "Worker 10":
				self.ws['C11'].value = "WORK"
			elif name == "Worker 11":
				self.ws['C12'].value = "WORK"
			elif name == "Worker 12":
				self.ws['C13'].value = "WORK"
			elif name == "Worker 13":
				self.ws['C14'].value = "WORK"
			else:
				pass


	def setup_excel(self):
		fontStyle = Font(size = "13")
		fontDefault = Font(size = "11")

		self.wb = load_workbook(self.config['location']['address'])
		ws1 = self.wb.create_sheet("pop")
		self.wb.active = self.wb["pop"]
		self.ws = self.wb.active
		self.ws.title = f"{self.sat_wkst_title_date} - {self.sun_wkst_title_date}" 
		self.ws.sheet_view.zoomScale = 150 # Set zoom to 150% for the worksheet we just created
		self.ws['A1'].alignment = Alignment(wrap_text=True)
		self.ws['B1'].alignment = Alignment(wrap_text=True)
		self.ws['C1'].alignment = Alignment(wrap_text=True)
		self.ws.cell(row = 1, column =1, value = 'Weekend Work').font = fontStyle
		self.ws.cell(row = 1, column =2, value = f'Saturday \n({self.saturday})').font = fontDefault
		self.ws.cell(row = 1, column =3, value = f'Sunday \n({self.sunday})').font = fontDefault
		self.ws.row_dimensions[1].height = 43.80
		self.ws.column_dimensions['A'].width = 16.90
		self.set_border(self.ws, 'A1:C14')
		self.set_off_days()
		self.populate_cells()
		self.set_working()


	def save_workbook(self):
		self.wb.save(self.config['location']['address'])


	def open_excel(self):
		file = self.config['location']['address']
		os.startfile(file)